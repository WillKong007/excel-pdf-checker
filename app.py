import streamlit as st
import os
import tempfile
import zipfile
import io
import re
import logging
import unicodedata
import time
from collections import defaultdict
import numpy as np
import pandas as pd
import pdfplumber
import cv2
import pytesseract
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from concurrent.futures import ProcessPoolExecutor, as_completed

# ------------------------------------------------------------
# 0. HELPERS
# ------------------------------------------------------------
def normalize_ocr(text):
    mapping = {
        'O': '0', 'o': '0', 'Q': '0', 'q': '0',
        'I': '1', 'l': '1', '|': '1', '!': '1',
        'S': '5', 's': '5',
        'Z': '2', 'z': '2',
        'G': '6', 'g': '6',
        'T': '7', 't': '7',
        'B': '8', 'b': '8',
        'A': '4', 'a': '4',
        'E': '3', 'e': '3',
    }
    for old, new in mapping.items():
        text = text.replace(old, new)
    return text

def safe_normalize(token):
    if re.search(r'\d', token):
        return normalize_ocr(token)
    return token

def process_single_page(pdf_path, page_num, dpi, ocr_config):
    try:
        images = convert_from_path(pdf_path, dpi=dpi, first_page=page_num, last_page=page_num)
        if not images:
            return page_num, ""
        img = np.array(images[0])
        gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
        thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
        text = pytesseract.image_to_string(thresh, lang='eng', config=ocr_config)
        return page_num, text
    except Exception as e:
        return page_num, f"ERROR: {e}"

# ------------------------------------------------------------
# 1. ALPHANUMERIC CHECKER - IMPROVED
# ------------------------------------------------------------
EN_CONFIG = {
    "missing_threshold": 0.05,
    "ignore_hidden": True,
    "use_print_area": False,
    "min_word_length": 2,
    "token_presence_threshold": 1.0,
    "force_ocr": True,
    "max_pages": 0,
    "ocr_dpi": 350,
    "parallel_workers": 2,
    "use_dual_engine": True,
    "use_substring_check": True,
    "ocr_psm": 6,
    "log_file": "alphanumeric_check.log"
}

class ExcelPDFAlphanumericChecker:
    def __init__(self, config=None):
        self.config = config or EN_CONFIG
        self.results = []
        self.ignore_tokens = {'x000f', 'x001a', 'x001b', '_x000f_'}

    def normalize_text(self, text):
        text = re.sub(r'[\x00-\x1f\x7f]', '', text)
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    def extract_excel_content(self, excel_path):
        wb = load_workbook(excel_path, data_only=True)
        all_text = []
        cell_map = {}
        col_counts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print_area = None
            if self.config["use_print_area"] and ws.print_area:
                if isinstance(ws.print_area, str):
                    print_area = ws.print_area.split(',')[0]
                else:
                    print_area = ws.print_area[0]
            if print_area:
                try:
                    if '!' in print_area:
                        print_area = print_area.split('!')[-1]
                    min_col_b, min_row_b, max_col_b, max_row_b = range_boundaries(print_area)
                    min_row = min_row_b or 1
                    max_row = max_row_b or ws.max_row
                    min_col = min_col_b or 1
                    max_col = max_col_b or ws.max_column
                except:
                    min_row, max_row = 1, ws.max_row
                    min_col, max_col = 1, ws.max_column
            else:
                min_row, max_row = 1, ws.max_row
                min_col, max_col = 1, ws.max_column
            col_counts.append(max_col - min_col + 1)
            for row in range(min_row, max_row + 1):
                if self.config["ignore_hidden"] and ws.row_dimensions[row].hidden:
                    continue
                for col in range(min_col, max_col + 1):
                    col_letter = get_column_letter(col)
                    if self.config["ignore_hidden"] and ws.column_dimensions[col_letter].hidden:
                        continue
                    cell = ws.cell(row=row, column=col)
                    value = cell.value
                    if value is None:
                        continue
                    text = str(value).strip()
                    if not text or len(text) < self.config["min_word_length"]:
                        continue
                    cell_ref = f"{col_letter}{row}"
                    all_text.append(text)
                    cell_map[(sheet_name, cell_ref)] = text
        wb.close()
        avg_cols = int(np.mean(col_counts)) if col_counts else 3
        return ' '.join(all_text), all_text, cell_map, avg_cols

    def extract_pdf_content_parallel(self, pdf_path, progress_callback=None):
        all_text = []
        max_pages = self.config.get("max_pages", 0)
        dpi = self.config.get("ocr_dpi", 350)
        psm = self.config.get("ocr_psm", 6)
        ocr_config = f'--psm {psm}'
        workers = self.config.get("parallel_workers", 2)
        if max_pages == 0:
            with pdfplumber.open(pdf_path) as pdf:
                total_pages = len(pdf.pages)
        else:
            total_pages = max_pages
        page_numbers = range(1, total_pages + 1)
        args_list = [(pdf_path, i, dpi, ocr_config) for i in page_numbers]
        with ProcessPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(process_single_page, *args): args[1] for args in args_list}
            completed = 0
            for future in as_completed(futures):
                page_num, text = future.result()
                if text:
                    all_text.append(text)
                completed += 1
                if progress_callback:
                    progress_callback(completed, total_pages)
        combined = re.sub(r'\s+', ' ', ' '.join(all_text)).strip()
        return combined, all_text

    def extract_pdf_content(self, pdf_path, progress_callback=None):
        all_text = []
        if self.config.get("use_dual_engine", False):
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    pages = pdf.pages[:self.config.get("max_pages")] if self.config.get("max_pages") > 0 else pdf.pages
                    for idx, page in enumerate(pages):
                        text = page.extract_text(layout=True)
                        if text:
                            all_text.append(text)
                        if progress_callback:
                            progress_callback(idx+1, len(pages))
            except Exception as e:
                logging.warning(f"pdfplumber failed: {e}")
            ocr_text, _ = self.extract_pdf_content_parallel(pdf_path)
            if ocr_text:
                all_text.append(ocr_text)
        else:
            # fallback logic (original)
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    pages = pdf.pages[:self.config.get("max_pages")] if self.config.get("max_pages") > 0 else pdf.pages
                    for idx, page in enumerate(pages):
                        text = page.extract_text(layout=True)
                        if text:
                            all_text.append(text)
                        if progress_callback:
                            progress_callback(idx+1, len(pages))
            except:
                pass
            if not all_text or len(''.join(all_text).strip()) < 100:
                ocr_text, _ = self.extract_pdf_content_parallel(pdf_path, progress_callback)
                if ocr_text:
                    all_text.append(ocr_text)
        combined = re.sub(r'\s+', ' ', ' '.join(all_text)).strip()
        return combined, all_text

    def tokenize_english(self, text):
        if not text:
            return []
        text = unicodedata.normalize('NFKC', text)
        tokens = re.findall(r'[a-zA-Z0-9]+(?:[-\u2010-\u2015][a-zA-Z0-9]+)*', text)
        return [t.lower() for t in tokens if len(t) >= self.config["min_word_length"]]

    def is_global_duplicate(self, token, cell_map, current_cell):
        token_lower = token.lower()
        for (s, r), val in cell_map.items():
            if (s, r) != current_cell and token_lower in str(val).lower():
                return True
        return False

    def detect_visual_truncation(self, cell_text, pdf_text):
        if len(cell_text) < 8:
            return False, ""
        cell_clean = re.sub(r'\s+', '', self.normalize_text(cell_text))
        pdf_clean = re.sub(r'\s+', '', self.normalize_text(pdf_text))
        if len(cell_clean) > 50 and len(cell_clean) / max(len(pdf_clean), 1) < 0.75:
            return True, "PARTIAL VISUAL TRUNCATION (length mismatch)"
        suffix = cell_clean[-min(40, len(cell_clean)//2):]
        if suffix and suffix not in pdf_clean:
            return True, f"PARTIAL VISUAL TRUNCATION: suffix '{suffix[-20:]}'"
        return False, ""

    def compare_content(self, excel_text, pdf_text, cell_map, avg_cols):
        threshold = 0.95 if avg_cols <= 3 else (0.91 if avg_cols <= 5 else 0.87)
        excel_tokens = set(self.tokenize_english(excel_text))
        pdf_tokens = set(self.tokenize_english(pdf_text))
        pdf_clean_sub = re.sub(r'[^a-zA-Z0-9]', '', re.sub(r'\s+', '', pdf_text.lower()))

        missing_cells = defaultdict(list)
        truncated_cells = {}

        for (sheet, ref), val in cell_map.items():
            tokens_in_cell = set(self.tokenize_english(val))
            if not tokens_in_cell:
                continue
            cell_clean = re.sub(r'[^a-zA-Z0-9]', '', re.sub(r'\s+', '', val))
            if cell_clean in pdf_clean_sub:
                continue

            is_trunc, trunc_msg = self.detect_visual_truncation(val, pdf_text)
            if is_trunc:
                truncated_cells[f"{sheet}!{ref}"] = trunc_msg
                continue

            present = tokens_in_cell & pdf_tokens
            ratio = len(present) / len(tokens_in_cell)
            if ratio < threshold:
                cell_missing = [t for t in (tokens_in_cell - pdf_tokens - self.ignore_tokens)
                                if not self.is_global_duplicate(t, cell_map, (sheet, ref))]
                if cell_missing:
                    missing_cells[(sheet, ref)].extend(cell_missing)

        try:
            tfidf = TfidfVectorizer()
            similarity = cosine_similarity(tfidf.fit_transform([excel_text, pdf_text]))[0][0]
        except:
            similarity = 0.0

        stats = {
            "excel_word_count": len(excel_tokens),
            "pdf_word_count": len(pdf_tokens),
            "missing_word_count": sum(len(v) for v in missing_cells.values()),
            "missing_percentage": sum(len(v) for v in missing_cells.values()) / max(len(excel_tokens), 1) * 100,
            "similarity_score": similarity,
            "missing_cells": dict(missing_cells),
            "truncated_cells": truncated_cells,
            "truncated_count": len(truncated_cells),
            "avg_columns": avg_cols,
            "pdf_text": pdf_text,
            "cell_map": cell_map,
        }
        return stats

    def check_pair(self, excel_path, pdf_path, progress_callback=None):
        excel_text, _, cell_map, avg_cols = self.extract_excel_content(excel_path)
        pdf_text, _ = self.extract_pdf_content(pdf_path, progress_callback)
        if not excel_text or not pdf_text:
            return None
        stats = self.compare_content(excel_text, pdf_text, cell_map, avg_cols)
        stats["excel_file"] = os.path.basename(excel_path)
        stats["pdf_file"] = os.path.basename(pdf_path)
        self.results.append(stats)
        return stats

    def generate_full_report(self):
        lines = ["=" * 70, "ALPHANUMERIC (ENGLISH & NUMBERS) LOSS DETECTION REPORT", "=" * 70]
        for res in self.results:
            lines.append(f"\n📄 {res['excel_file']} ↔ {res['pdf_file']}")
            lines.append(f"   Columns: {res.get('avg_columns', 3)} | Similarity: {res['similarity_score']:.4f}")
            if res.get('truncated_cells'):
                lines.append("   🚨 Truncated cells:")
                for c, m in res['truncated_cells'].items():
                    lines.append(f"     {c} → {m}")
            if res['missing_cells']:
                lines.append("   Missing content cells:")
                for (s, r), toks in res['missing_cells'].items():
                    lines.append(f"     {s}!{r}: {', '.join(toks)}")
        lines.append("=" * 70)
        return "\n".join(lines)

# ------------------------------------------------------------
# 2. CHINESE CHECKER - IMPROVED
# ------------------------------------------------------------
ZH_CONFIG = {
    "missing_threshold": 0.02,
    "ignore_hidden": True,
    "use_print_area": True,
    "cell_presence_threshold": 0.95,
    "ocr_psm_zh": 6,
}

class ExcelPDFChineseChecker:
    def __init__(self, config=None):
        self.config = config or ZH_CONFIG
        self.results = []

    def ultimate_normalize(self, text):
        return unicodedata.normalize('NFKC', str(text).replace('_x000F_', '').strip())

    def extract_excel_content(self, excel_path):
        wb = load_workbook(excel_path, data_only=True)
        all_text = []
        cell_map = {}
        col_counts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            min_row, max_row = 1, ws.max_row
            min_col, max_col = 1, ws.max_column
            if self.config["use_print_area"] and ws.print_area:
                try:
                    print_area = ws.print_area[0] if isinstance(ws.print_area, list) else ws.print_area.split(',')[0]
                    if '!' in print_area:
                        print_area = print_area.split('!')[-1]
                    min_col_b, min_row_b, max_col_b, max_row_b = range_boundaries(print_area)
                    min_row = min_row_b or 1
                    max_row = max_row_b or ws.max_row
                    min_col = min_col_b or 1
                    max_col = max_col_b or ws.max_column
                except:
                    pass
            col_counts.append(max_col - min_col + 1)
            for row in range(min_row, max_row + 1):
                if self.config["ignore_hidden"] and ws.row_dimensions[row].hidden: continue
                for col in range(min_col, max_col + 1):
                    col_letter = get_column_letter(col)
                    if self.config["ignore_hidden"] and ws.column_dimensions[col_letter].hidden: continue
                    value = ws.cell(row=row, column=col).value
                    if not value: continue
                    text = self.ultimate_normalize(value)
                    if not re.search(r'[\u4e00-\u9fff]', text): continue
                    cell_ref = f"{col_letter}{row}"
                    all_text.append(text)
                    cell_map[(sheet_name, cell_ref)] = text
        wb.close()
        avg_cols = int(np.mean(col_counts)) if col_counts else 3
        return ' '.join(all_text), all_text, cell_map, avg_cols

    def extract_pdf_content(self, pdf_path, progress_callback=None):
        all_text = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if text:
                        all_text.append(self.ultimate_normalize(text))
                    if progress_callback:
                        progress_callback(idx+1, len(pdf.pages))
        except:
            pass
        try:
            images = convert_from_path(pdf_path, dpi=200)
            psm = self.config.get("ocr_psm_zh", 6)
            for idx, img in enumerate(images):
                gray = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2GRAY)
                _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)
                ocr_text = pytesseract.image_to_string(thresh, lang='chi_tra', config=f'--psm {psm}')
                if ocr_text:
                    all_text.append(self.ultimate_normalize(ocr_text))
                if progress_callback:
                    progress_callback(idx+1, len(images))
        except:
            pass
        return ' '.join(all_text), all_text

    def tokenize_chinese(self, text):
        return re.findall(r'[\u4e00-\u9fff]', self.ultimate_normalize(text))

    def check_cell_presence(self, cell_text, pdf_text, avg_cols):
        cell_chars = self.tokenize_chinese(cell_text)
        if not cell_chars:
            return True, ""
        pdf_stream = "".join(self.tokenize_chinese(pdf_text))
        cell_str = "".join(cell_chars)
        if cell_str in pdf_stream:
            return True, ""
        ngram_thresh = 0.03 if len(cell_chars) < 20 else 0.09
        # dynamic suffix check etc.
        suffix_len = min(6, len(cell_chars))
        if "".join(cell_chars[-suffix_len:]) not in pdf_stream and len(cell_chars) > 10:
            return False, "結尾被裁切"
        return True, ""

    def compare_content(self, excel_text, pdf_text, cell_map, avg_cols):
        missing_cells = defaultdict(list)
        truncated_cells = {}
        for (sheet, ref), val in cell_map.items():
            is_present, msg = self.check_cell_presence(val, pdf_text, avg_cols)
            if not is_present:
                truncated_cells[f"{sheet}!{ref}"] = msg
        stats = {
            "missing_cells": dict(missing_cells),
            "truncated_cells": truncated_cells,
            "truncated_count": len(truncated_cells),
            "avg_columns": avg_cols,
            "pdf_text": pdf_text,
            "cell_map": cell_map,
        }
        return stats

    def check_pair(self, excel_path, pdf_path, progress_callback=None):
        excel_text, _, cell_map, avg_cols = self.extract_excel_content(excel_path)
        pdf_text, _ = self.extract_pdf_content(pdf_path, progress_callback)
        stats = self.compare_content(excel_text, pdf_text, cell_map, avg_cols)
        stats["excel_file"] = os.path.basename(excel_path)
        stats["pdf_file"] = os.path.basename(pdf_path)
        self.results.append(stats)
        return stats

    def generate_report(self):
        lines = ["=" * 70, "中文內容漏字與裁切審查報告", "=" * 70]
        for res in self.results:
            lines.append(f"\n📄 {res['excel_file']} ↔ {res['pdf_file']}")
            if res.get('truncated_cells'):
                lines.append("   截斷儲存格：")
                for c, m in res['truncated_cells'].items():
                    lines.append(f"     {c} → {m}")
        lines.append("=" * 70)
        return "\n".join(lines)

# ------------------------------------------------------------
# 3. RUNNER & HELPERS
# ------------------------------------------------------------
def run_check(excel_path, pdf_path, mode='both', progress_callback=None):
    results = {}
    if mode in ('english', 'both'):
        checker = ExcelPDFAlphanumericChecker()
        stats = checker.check_pair(excel_path, pdf_path, progress_callback)
        if stats:
            results['alphanumeric'] = {'stats': stats, 'report': checker.generate_full_report()}
    if mode in ('chinese', 'both'):
        checker = ExcelPDFChineseChecker()
        stats = checker.check_pair(excel_path, pdf_path, progress_callback)
        if stats:
            results['chinese'] = {'stats': stats, 'report': checker.generate_report()}
    return results

def process_zip(zip_file, mode='both'):
    results = []
    with zipfile.ZipFile(zip_file, 'r') as z:
        with tempfile.TemporaryDirectory() as tmpdir:
            z.extractall(tmpdir)
            files = os.listdir(tmpdir)
            excel_files = [f for f in files if f.endswith('.xlsx')]
            for ex in excel_files:
                base = os.path.splitext(ex)[0]
                pdfs = [f for f in files if f.endswith('.pdf') and os.path.splitext(f)[0] == base]
                if pdfs:
                    result = run_check(os.path.join(tmpdir, ex), os.path.join(tmpdir, pdfs[0]), mode)
                    if result:
                        results.append({"excel": ex, "pdf": pdfs[0], "results": result})
    return results

def generate_excel_report(all_results):
    rows = []
    for lang, data in all_results.items():
        stats = data['stats']
        for (sheet, ref), tokens in stats.get('missing_cells', {}).items():
            rows.append({"Language": lang, "Sheet": sheet, "Cell": ref, "Missing": ", ".join(tokens)})
        for cell, msg in stats.get('truncated_cells', {}).items():
            sheet, ref = cell.split("!") if "!" in cell else ("Unknown", cell)
            rows.append({"Language": lang, "Sheet": sheet, "Cell": ref, "Missing": msg})
    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    return output.getvalue()

# ------------------------------------------------------------
# 4. STREAMLIT UI
# ------------------------------------------------------------
st.set_page_config(page_title="Excel ↔ PDF Missing Content Checker", layout="wide")
st.title("📄 Excel ↔ PDF Content Loss Detector")
st.markdown("**Improved version**: Multi-column support • Visual truncation detection • Adaptive thresholds")

mode = st.radio("Select mode:", ["Both", "Alphanumeric only", "Chinese only"], index=0)
mode_map = {"Both": "both", "Alphanumeric only": "english", "Chinese only": "chinese"}

upload_type = st.radio("Upload type:", ["Single File", "Batch ZIP"], index=0)

if upload_type == "Single File":
    col1, col2 = st.columns(2)
    with col1:
        excel_file = st.file_uploader("Excel (.xlsx)", type=["xlsx"])
    with col2:
        pdf_file = st.file_uploader("PDF (.pdf)", type=["pdf"])
else:
    zip_file = st.file_uploader("ZIP file", type=["zip"])

if st.button("🔍 Start Check"):
    if upload_type == "Single File":
        if not excel_file or not pdf_file:
            st.error("Please upload both files.")
            st.stop()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_ex:
            tmp_ex.write(excel_file.read())
            ex_path = tmp_ex.name
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
            tmp_pdf.write(pdf_file.read())
            pdf_path = tmp_pdf.name

        status = st.status("Analyzing...", expanded=True)
        progress_bar = st.progress(0)
        def cb(c, t): 
            progress_bar.progress(c / t)
        results = run_check(ex_path, pdf_path, mode_map[mode], cb)
        status.update(label="✅ Done", state="complete")
        os.unlink(ex_path)
        os.unlink(pdf_path)
    else:
        if not zip_file:
            st.error("Upload ZIP")
            st.stop()
        results_list = process_zip(zip_file, mode_map[mode])
        results = {}
        for item in results_list:
            for k, v in item["results"].items():
                results[k] = v

    # Display results (same as your original UI)
    if results:
        st.success("Analysis complete!")
        for lang, data in results.items():
            st.subheader(f"{lang.upper()} Results")
            stats = data['stats']
            st.text_area("Report", data['report'], height=300)
            if stats.get('missing_cells') or stats.get('truncated_cells'):
                st.dataframe(pd.DataFrame([
                    {"Cell": k, "Issue": v} for k, v in {**stats.get('missing_cells', {}), **stats.get('truncated_cells', {})}.items()
                ]))

        # Download buttons
        if any(r.get('stats', {}).get('missing_cells') for r in results.values()):
            st.download_button("Download Excel Report", generate_excel_report(results), "report.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.download_button("Download Full Report", "\n\n".join(d['report'] for d in results.values()), "full_report.txt")

st.caption("Ready for Streamlit Cloud • Improved for any number of columns + visual truncation")
